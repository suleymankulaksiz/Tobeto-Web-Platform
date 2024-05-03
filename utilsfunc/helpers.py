import json
import time
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from constants.globalConstants import *
from selenium.webdriver.common.action_chains import ActionChains
from PIL import Image
from selenium.webdriver.support.ui import Select
import time




def readInvalidDataFromExcel():
        excelFile= openpyxl.load_workbook("data/invalid_Login.xlsx")
        sheet= excelFile["Sayfa1"]
        data= []
        for rows in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
            loginEmailx, loginPasswordx =rows
            data.append((loginEmailx, loginPasswordx))
        return data


def readInvalidDataFromJSON():
        with open(r'data/data.json', 'r') as file: # ilk baştaki r ön ek # testlerde genelde r kullanılır yani dosyayı okuma modunda açar.
            data = json.load(file) #JSON veriler Python veri yapısına dönüşür.
        return [(user['firstnamex'], user['lastnamex'], user['emailx'], user['passwordx'], user['passwordagainx']) for user in data['invalid_login_users']] #tuple list oluşturur.





class Helpfunc:
   
    def waitForElementVisible(self, locator, timeout=10):
        return WebDriverWait(self.driver, timeout).until(EC.visibility_of_element_located(locator))
    def waitForElementVisible(self, locator, timeout=5):
        return WebDriverWait(self.driver, timeout).until(EC.visibility_of_element_located(locator)) 
    def take_and_show_screenshot(self, image_path):
        screenshot = Image.open(image_path)
        screenshot.show()
    def checkboxs(self):
        checkbox1=self.waitForElementVisible((By.XPATH,CHECKBOX1_XPATH))
        checkbox1.click()
        checkbox2=self.waitForElementVisible((By.XPATH,CHECKBOX2_XPATH))
        checkbox2.click()
        checkbox3=self.waitForElementVisible((By.XPATH,CHECKBOX3_XPATH))
        checkbox3.click()
        checkbox4=self.waitForElementVisible((By.XPATH,CHECKBOX4_XPATH))
        checkbox4.click()
    def registerForm(self):
        emailrandom = generate_random_email()
        firstname = self.waitForElementVisible((By.NAME, FIRSTNAME_NAME))
        lastname = self.waitForElementVisible((By.NAME, LASTNAME_NAME))
        email = self.waitForElementVisible((By.NAME, EMAIL_NAME))
        password = self.waitForElementVisible((By.NAME, PASSWORD_NAME))
        password_again = self.waitForElementVisible((By.NAME, PASSWORDAGAIN_NAME))
        firstname.send_keys(input_firstname)
        lastname.send_keys(input_lastname)
        email.send_keys(emailrandom)
        password.send_keys(input_password)
        password_again.send_keys(input_passwordagain)
        sign_up_button = self.waitForElementVisible((By.XPATH, SIGNUPBUTTON_XPATH))
        sign_up_button.click()
    def phoneContracts(self):
        phone_checkbox=self.waitForElementVisible((By.XPATH,PHONECHECK_XPATH))
        phone_checkbox.click()
        input_phone_box=self.waitForElementVisible((By.XPATH,PHONECHECK_XPATH))
        input_phone_box.send_keys(input_phone)
    def country_select(self):
           excelFile = openpyxl.load_workbook('data/countryid_list.xlsx')
           sheet = excelFile.active
           country_ids = [cell.value for cell in sheet['A'][1:]]  # İlk satırı atla (başlık)
           de_ids = [country_id for country_id in country_ids if country_id == 'TR']
           if not de_ids:
             print("TR değeri bulunamadı.")
             excelFile.close()
             return
           
           phone_country_field = self.driver.find_element(By.XPATH, "/html/body/div[4]/div/div/div/div/div/label[4]/small/div/div/select")
           select = Select(phone_country_field)
           select.select_by_value(de_ids[0])
           
           
    def precondition(self):
        forgot_password_button = self.waitForElementVisible((By.XPATH,FORGOT_PASSWORD_XPATH))
        forgot_password_button.click()
        forgot_email = self.waitForElementVisible((By.XPATH, FORGOT_EMAIL_XPATH))
        forgot_email.send_keys(input_forgot_email)
        sendButton = self.waitForElementVisible((By.XPATH,SENDBUTTON_XPATH))
        sendButton.click()

        popupMessage = self.waitForElementVisible((By.XPATH,FORGOT_EMAIL_POPUP_XPATH))
        assert popupMessage.text == FORGOT_EMAIL_POPUP_TEXT
        time.sleep(2)
        alert_quit = self.waitForElementVisible((By.XPATH, FORGOT_EMAIL_POPUP_XPATH))
        alert_quit.click()

        self.driver.get(SIGN_IN)    
        sign_in_email = self.waitForElementVisible((By.ID,SIGNINEMAIL_ID))
        sign_in_email.send_keys(input_forgot_email)
        next_button = self.waitForElementVisible((By.XPATH,LOGINNEXTBUTTON_XPATH))
        next_button.click()
        sign_in_password = self.waitForElementVisible((By.XPATH,SINGINPASSWORD_XPATH))
        sign_in_password.send_keys(input_sign_in_password)
        sign_in_button = self.waitForElementVisible((By.CSS_SELECTOR,LOGINSINGINBUTTON_XPATH))
        sign_in_button.click()
        #Last mail click
        last_mail_link = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, LASTMAILLINK_CSS))) 
        last_mail_link.click()
        time.sleep(4)
        #Açılan sayfaya geçer
        windows = self.driver.window_handles
        self.driver.switch_to.window(windows[-1])
        #Mail içerisindeki linke tıklar.
        email_content_xpath = EMAILCONTENT_XPATH
        email_content_element = self.waitForElementVisible((By.XPATH,email_content_xpath))
        email_content = email_content_element.text
        link_start_index = email_content.find("https://")  # Linkin başlangıç indeksi
        link_end_index = email_content.find(" ", link_start_index)  # Linkin sonraki boşluk karakterine kadar olan kısmı alır

        if link_start_index != -1:
            link_url = email_content[link_start_index:link_end_index]  # Linki al
            # Linki tarayıcıda aç
            self.driver.get(link_url)
            
            
    def homepage_btn_click(self):
        homepage_nav_button=self.waitForElementVisible((By.XPATH, HOMEPAGE_NAV_XPATH))
        homepage_nav_button.click()
        time.sleep(2)
        self.driver.execute_script("window.scrollTo(0,1000)")
        time.sleep(2)
        
    def waitForElementVisible(self, locator, timeout=10):
        return WebDriverWait(self.driver, timeout).until(EC.visibility_of_element_located(locator)) 
    def js_click(self, element):
        self.driver.execute_script("arguments[0].click();", element)
    def login(self):
        loginEmail = self.waitForElementVisible((By.NAME, EMAIL_NAME))
        loginEmail.send_keys(input_setting_email)
        loginPassword = self.waitForElementVisible((By.NAME, PASSWORD_NAME))
        loginPassword.send_keys(input_setting_password)
        loginButton = self.waitForElementVisible((By.XPATH, LOGIN_BUTTON_XPATH))
        loginButton.click()
        alert_quit = self.waitForElementVisible((By.XPATH, SETTING_LOGIN_ALERT_XPATH))
        alert_quit.click()
        wait_url = PLATFORM_URL
        WebDriverWait(self.driver, 10).until(EC.url_to_be(wait_url))
        create_profile_button = self.waitForElementVisible((By.XPATH, SET_PROFILE_BUTTON_XPATH))
        self.js_click(create_profile_button)
        setting_button=self.waitForElementVisible((By.XPATH,SETTING_BUTTON_XPATH))
        setting_button.click()
    def login_two(self):
         loginEmail = self.waitForElementVisible((By.NAME, EMAIL_NAME))
         loginEmail.send_keys(input_setting_email_two)
         loginPassword = self.waitForElementVisible((By.NAME, PASSWORD_NAME))
         loginPassword.send_keys(input_setting_password_two)
         loginButton = self.waitForElementVisible((By.XPATH, LOGIN_BUTTON_XPATH))
         loginButton.click()
         alert_quit = self.waitForElementVisible((By.XPATH, SETTING_LOGIN_ALERT_XPATH))
         alert_quit.click()
         wait_url = PLATFORM_URL
         WebDriverWait(self.driver, 10).until(EC.url_to_be(wait_url))
         create_profile_button = self.waitForElementVisible((By.XPATH, SET_PROFILE_BUTTON_XPATH))
         self.js_click(create_profile_button)
         setting_button=self.waitForElementVisible((By.XPATH,SETTING_BUTTON_XPATH))
         setting_button.click()