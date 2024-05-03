import pytest
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from constants.globalConstants import *
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains


class Test_Login():
    @pytest.fixture(autouse=True)
    def setup(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(LOGIN_URL)
        yield
        self.driver.quit()
    
    def waitForElementVisible(self, locator, timeout=5):
        return WebDriverWait(self.driver, timeout).until(EC.visibility_of_element_located(locator))
    

    #Kullanıcının giriş yapma işlemi
    def test_login(self):
        loginEmail= self.waitForElementVisible((By.NAME, loginEmail_name))
        loginEmail.send_keys(input_loginEmail)
        loginPassword = self.waitForElementVisible((By.NAME, loginPassword_name))
        loginPassword.send_keys(input_loginPassword)
        loginButton= self.waitForElementVisible((By.XPATH, loginButton_xpath))
        loginButton.click()
        WebDriverWait(self.driver,5).until(EC.url_to_be(HOMEPAGE_URL))
        assert HOMEPAGE_URL in self.driver.current_url
        
    #Kullanıcının doldurulması zorunlu alanları boş bırakması
    def test_empty_fields(self):
        loginEmail= self.waitForElementVisible((By.NAME, loginEmail_name))
        loginEmail.send_keys(input_empty)
        loginPassword = self.waitForElementVisible((By.NAME, loginPassword_name))
        loginPassword.send_keys(input_empty)
        loginButton= self.waitForElementVisible((By.XPATH, loginButton_xpath))
        loginButton.click()
        emptyEmailError= self.waitForElementVisible((By.XPATH, loginEmptyEmail_xpath))
        emptyPasswordError= self.waitForElementVisible((By.XPATH, loginEmptyPassword_xpath))
        assert emptyEmailError.text == empty_fields_text and emptyPasswordError.text == empty_fields_text

    
    #Geçersiz e-posta veya şifre ile giriş yapma işlemi
    def readInvalidDataFromExcel():
        excelFile= openpyxl.load_workbook("data/invalid_Login.xlsx")
        sheet= excelFile["Sayfa1"]
        data= []
        for rows in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
            loginEmailx, loginPasswordx =rows
            data.append((loginEmailx, loginPasswordx))
        return data
    
    @pytest.mark.parametrize("loginEmailx, loginPasswordx", readInvalidDataFromExcel())
    def test_invalid_login(self, loginEmailx, loginPasswordx):
        loginEmail= self.waitForElementVisible((By.NAME, loginEmail_name))
        loginEmail.send_keys(loginEmailx)
        loginPassword = self.waitForElementVisible((By.NAME, loginPassword_name))
        loginPassword.send_keys(loginPasswordx)
        loginButton= self.waitForElementVisible((By.XPATH, loginButton_xpath))
        loginButton.click()
        invalidLoginError = self.waitForElementVisible((By.XPATH, invalidLoginError_xpath))
        assert invalidLoginError.text == login_invalidLogin_text

    #kullanıcının kayıt ol sayfasına yönlendirilmesi
    def test_sign_up(self):
        signUpButton= self.waitForElementVisible((By.XPATH, loginSignUp_xpath))
        signUpButton.click()
        assert REGISTER_URL in self.driver.current_url
        
    #Doğrulanmayan e-posta ile sisteme giriş yapılması
    def test_unverified_login(self):
        loginEmail= self.waitForElementVisible((By.NAME, loginEmail_name))
        loginEmail.send_keys(input_unverified_email)
        loginPassword = self.waitForElementVisible((By.NAME, loginPassword_name))
        loginPassword.send_keys(input_unverified_password)
        loginButton= self.waitForElementVisible((By.XPATH, loginButton_xpath))
        loginButton.click()
        unverifiedLoginError = self.waitForElementVisible((By.XPATH, unverifiedLoginError_xpath))
        assert unverifiedLoginError.text == login_unverifiedLogin_text
        
        

    


